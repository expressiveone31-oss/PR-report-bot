"""
Модуль обновления охватов в xlsx медиаплане.

Алгоритм:
1. Читает xlsx, находит колонки со ссылками на публикации и «Охват (факт)»
2. По каждой ссылке идёт в API нужной платформы
3. Проставляет фактический охват в ячейку
4. Возвращает обновлённый xlsx как bytes
"""

import io
import re
import asyncio
import logging
from typing import Optional

import openpyxl
from openpyxl.styles import PatternFill

logger = logging.getLogger(__name__)

# Ключевые слова для поиска нужных колонок
POST_URL_KEYWORDS = ("ссылка на публикацию", "ссылка на пост", "ссылка на твит",
                     "ссылка на рекламу", "публикация", "post url")
# Порядок важен: более точные совпадения первыми
# НЕ включаем просто "охват" — это подхватит "Планируемый охват"
REACH_FACT_KEYWORDS = ("охват (факт)", "охват факт", "просмотры факт", "просмотры (факт)",
                       "реальный охват", "итого охват", "итого просмотры",
                       "реальные просмотры", "факт охват", "факт просмотры", "views fact")

# Цвета ячеек
UPDATED_FILL = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")   # жёлтый — охват обновлён
ERROR_FILL   = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")   # красный — нет данных
DELETED_FILL = PatternFill(start_color="FFD0E4", end_color="FFD0E4", fill_type="solid")   # розовый — пост удалён


def _detect_platform(url: str) -> str:
    url = url.lower()
    if "vk.com" in url or "vk.ru" in url:
        return "vk"
    if "t.me" in url or "telegram" in url:
        return "telegram"
    if "instagram.com" in url:
        return "instagram"
    if "youtube.com" in url or "youtu.be" in url:
        return "youtube"
    if "tiktok.com" in url or "vt.tiktok.com" in url:
        return "tiktok"
    if "x.com" in url or "twitter.com" in url:
        return "twitter"
    return "unknown"


def _is_post_url(url: str) -> bool:
    """Проверяет что ссылка ведёт на конкретный пост, а не на канал."""
    url = url.strip().lower()
    if not url.startswith("http"):
        return False
    # Исключаем ссылки на каналы/профили без ID поста
    skip_patterns = (
        "disk.yandex", "yandex.ru/i/", "prnt.sc", "drive.google",
        "clck.ru", "yandex.go",
    )
    if any(p in url for p in skip_patterns):
        return False
    # VK wall — должен содержать _
    if ("vk.com" in url or "vk.ru" in url):
        return "wall" in url or "clip" in url
    # Telegram — должен содержать /число в конце
    if "t.me" in url:
        return bool(re.search(r"/\d+$", url))
    # Instagram — reel или /p/
    if "instagram.com" in url:
        return "/reel/" in url or "/p/" in url
    # YouTube shorts
    if "youtube.com" in url or "youtu.be" in url:
        return "/shorts/" in url or "watch?v=" in url or "youtu.be/" in url
    # TikTok video
    if "tiktok.com" in url:
        return "/video/" in url or "vt.tiktok" in url
    # Twitter status
    if "x.com" in url or "twitter.com" in url:
        return "/status/" in url
    return False


def _find_col(headers: list[str], keywords: tuple,
              exclude_keywords: tuple = ()) -> Optional[int]:
    """Ищет колонку по ключевым словам в заголовке.
    exclude_keywords — слова которые НЕ должны быть в заголовке.
    """
    for i, h in enumerate(headers):
        h_lower = h.strip().lower()
        # Пропускаем если заголовок содержит слова-исключения
        if any(ex in h_lower for ex in exclude_keywords):
            continue
        for kw in keywords:
            if kw in h_lower:
                return i
    return None


async def _get_views(url: str) -> tuple[Optional[int], str]:
    """
    Идёт в нужный API и возвращает (просмотры, статус).
    Статус: "ok" | "deleted" | "error"
    """
    platform = _detect_platform(url)
    try:
        if platform == "vk":
            from src.platforms.vk import get_post_stats
            result = await get_post_stats(url)
            if result.error:
                return None, "error"
            return result.views, "ok"

        elif platform == "telegram":
            from src.platforms import telemetr, tgstat
            result = await telemetr.get_post_stats(url)
            telemetr_views = result.views or 0
            fallback = await tgstat.get_post_stats(url)
            tgstat_views = fallback.views or 0

            # Пост удалён: TGStat не находит, но Telemetr отдаёт кеш
            post_deleted = fallback.error in ("post_not_found_in_channel", "post_not_found")
            if post_deleted and telemetr_views > 0:
                return telemetr_views, "deleted"
            if post_deleted:
                return None, "error"

            if not fallback.error and tgstat_views > telemetr_views:
                return tgstat_views, "ok"
            return (telemetr_views, "ok") if telemetr_views > 0 else (None, "error")

        elif platform == "instagram":
            from src.platforms.hikerapi import get_post_stats
            result = await get_post_stats(url)
            if result.error:
                return None, "error"
            return result.views, "ok"

        elif platform == "youtube":
            from src.platforms.youtube import get_post_stats
            result = await get_post_stats(url)
            if result.error:
                return None, "error"
            return result.views, "ok"

        elif platform == "tiktok":
            from src.platforms.tiktok import get_post_stats
            result = await get_post_stats(url)
            if result.error:
                return None, "error"
            return result.views, "ok"

        elif platform == "twitter":
            from src.platforms.twitter import get_post_stats
            result = await get_post_stats(url)
            if result.error:
                return None, "error"
            return result.views, "ok"

    except Exception as e:
        logger.warning(f"API error for {url}: {e}")

    return None, "error"


async def update_xlsx(xlsx_bytes: bytes) -> tuple[bytes, dict]:
    """
    Обновляет охваты в xlsx.
    Возвращает (updated_xlsx_bytes, stats).
    stats = {"updated": N, "skipped": N, "errors": N}
    """
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    stats = {"updated": 0, "skipped": 0, "errors": 0, "deleted": 0}

    for sheet in wb.worksheets:
        # Ищем строку-заголовок
        header_row_idx = None
        headers = []
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
            row_strs = [str(c).strip() if c is not None else "" for c in row]
            row_lower = " ".join(row_strs).lower()
            if any(kw in row_lower for kw in ("ссылка на публикацию", "ссылка на пост",
                                               "ссылка на твит", "охват (факт)",
                                               "просмотры факт", "публикация",
                                               "охват факт", "реальный охват")):
                header_row_idx = row_idx
                headers = row_strs
                break

        if header_row_idx is None:
            logger.info(f"Sheet '{sheet.title}': no header found, skipping")
            continue

        col_url = _find_col(headers, POST_URL_KEYWORDS,
                            exclude_keywords=("канал", "channel", "профил"))
        col_reach = _find_col(headers, REACH_FACT_KEYWORDS,
                              exclude_keywords=("план", "прогноз", "ожидаем", "plan"))

        if col_url is None or col_reach is None:
            logger.warning(f"Sheet '{sheet.title}': col_url={col_url}, col_reach={col_reach} — не найдены нужные колонки")
            logger.warning(f"Sheet '{sheet.title}': headers={headers}")
            continue

        logger.info(f"Sheet '{sheet.title}': url_col={col_url} ('{headers[col_url]}'), reach_col={col_reach} ('{headers[col_reach]}')")

        # Собираем все URL для обработки.
        # Логика остановки:
        # 1. На "Итого:" — приостанавливаемся и проверяем: есть ли ниже органика?
        # 2. Если в ближайших 5 строках после "Итого:" встречается "Органика" или
        #    повторный заголовок "Ссылка на публикацию" — продолжаем сбор URL.
        # 3. Финальные стоп-слова: "Итого с органикой", "Сумма с НДС", "Общий прогнозируемый".
        FINAL_STOP_KEYWORDS = (
            "итого с органикой", "итого с органики", "сумма с учетом ндс",
            "сумма с ндс", "общий прогнозируемый", "общий прогноз",
            "фактический общий охват", "стоимость размещений",
        )
        ORGANIC_MARKERS = ("органика", "ссылка на публикацию", "ссылка на пост")

        url_rows = []
        row_idx = header_row_idx + 1
        max_row = sheet.max_row

        while row_idx <= max_row:
            # Собираем содержимое всей строки в одну строку для проверки маркеров
            row_text_parts = []
            for col_check in range(1, min(sheet.max_column, 15) + 1):
                cell_val = str(sheet.cell(row=row_idx, column=col_check).value or "").strip()
                if cell_val:
                    row_text_parts.append(cell_val.lower())
            row_text = " ".join(row_text_parts)

            # 1. Проверяем финальные стоп-слова — если есть, выходим
            if any(kw in row_text for kw in FINAL_STOP_KEYWORDS):
                logger.info(f"Updater: hit final stop at row {row_idx}: '{row_text[:80]}'")
                break

            # 2. Проверяем промежуточный "Итого:" (не финальный)
            is_intermediate_total = False
            for col_check in range(1, 4):
                cell_val = str(sheet.cell(row=row_idx, column=col_check).value or "").strip().lower()
                # Триггер "Итого" без "с органикой" и без "общий"
                if (cell_val.startswith("итого") and "органик" not in cell_val) or cell_val.startswith("итого (охват)"):
                    is_intermediate_total = True
                    break

            if is_intermediate_total:
                # Заглядываем в следующие 5 строк — есть ли там органика?
                organic_found = False
                for look_ahead in range(1, 6):
                    check_idx = row_idx + look_ahead
                    if check_idx > max_row:
                        break
                    check_row_parts = []
                    for col_check in range(1, min(sheet.max_column, 15) + 1):
                        cv = str(sheet.cell(row=check_idx, column=col_check).value or "").strip().lower()
                        if cv:
                            check_row_parts.append(cv)
                    check_text = " ".join(check_row_parts)
                    if any(m in check_text for m in ORGANIC_MARKERS):
                        organic_found = True
                        logger.info(f"Updater: organic block detected after 'Итого' at row {row_idx} (marker at row {check_idx})")
                        break

                if not organic_found:
                    # Органики нет — останавливаемся как раньше
                    logger.info(f"Updater: hit 'Итого' at row {row_idx}, no organic below — stopping")
                    break
                # Органика есть — продолжаем сканировать
                row_idx += 1
                continue

            # 3. Проверяем ячейку с URL
            cell_url = sheet.cell(row=row_idx, column=col_url + 1)
            url = str(cell_url.value).strip() if cell_url.value else ""
            if _is_post_url(url):
                url_rows.append((row_idx, url))

            row_idx += 1

        if not url_rows:
            continue

        # VK — последовательно с паузой
        vk_rows = [(r, u) for r, u in url_rows if _detect_platform(u) == "vk"]
        other_rows = [(r, u) for r, u in url_rows if _detect_platform(u) != "vk"]

        # Получаем просмотры
        views_map: dict[int, tuple[Optional[int], str]] = {}

        for row_idx, url in vk_rows:
            result = await _get_views(url)
            views_map[row_idx] = result
            await asyncio.sleep(0.4)

        if other_rows:
            results = await asyncio.gather(*[_get_views(u) for _, u in other_rows])
            for (row_idx, _), result in zip(other_rows, results):
                views_map[row_idx] = result

        # Проставляем в таблицу.
        # number_format сбрасываем в 'General' чтобы вставить число "как есть"
        # без наследуемого формата (₽, тысячные пробелы, копейки и т.п.)
        for row_idx, url in url_rows:
            views, status = views_map.get(row_idx, (None, "error"))
            cell_reach = sheet.cell(row=row_idx, column=col_reach + 1)

            if status == "ok" and views is not None:
                cell_reach.value = views
                cell_reach.number_format = 'General'
                cell_reach.fill = UPDATED_FILL
                stats["updated"] += 1
                logger.info(f"Updated row {row_idx}: {views:,} for {url}")
            elif status == "deleted" and views is not None:
                # Пост удалён — проставляем последний известный охват, розовый цвет
                cell_reach.value = views
                cell_reach.number_format = 'General'
                cell_reach.fill = DELETED_FILL
                stats["deleted"] += 1
                logger.info(f"Deleted post row {row_idx}: last known {views:,} for {url}")
            else:
                # Нет данных — оставляем пустым, красный цвет
                cell_reach.value = None
                cell_reach.fill = ERROR_FILL
                stats["errors"] += 1
                logger.warning(f"No data for row {row_idx}: {url}")

    # Сохраняем
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue(), stats
